from openpyxl import load_workbook

class ExcelProcessor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(filename=file_path, data_only=True)
        
    def _format_cell(self, value):
        """Format cell value to string, handling different data types."""
        if value is None:
            return 'None'
        elif isinstance(value, (int, float)):
            # Remove trailing zeros for floats
            return str(value).rstrip('0').rstrip('.') if '.' in str(value) else str(value)
        else:
            return str(value).strip()

    def _format_sheet_as_table(self, sheet):
        """Convert sheet data into a formatted string representation."""
        content = []
        
        print(f"min_row: {sheet.min_row}, max_row: {sheet.max_row}, min_col: {sheet.min_column}, max_col: {sheet.max_column}")
        content.append(f"=== Sheet: {sheet.title} ===")

        # Process each row
        for row in range(sheet.min_row, sheet.max_row + 1):
            row_values = []
            is_empty_row = True
            col = sheet.min_column
            
            while col <= sheet.max_column:
                cell = sheet.cell(row=row, column=col)
                value = self._format_cell(cell.value)
                if value != 'None':
                    is_empty_row = False
                row_values.append(value)
                col += 1
            
            # Only add non-empty rows
            if not is_empty_row:
                content.append('|' + '|'.join(row_values) + '|')

        return '\n'.join(content)

    def process_file(self):
        """Process entire workbook and return formatted content."""
        content = []
        
        # Process each sheet
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_content = self._format_sheet_as_table(sheet)
            if sheet_content.strip():  # Only add non-empty sheets
                content.append(sheet_content)
        
        return '\n'.join(content)
